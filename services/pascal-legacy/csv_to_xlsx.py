#!/usr/bin/env python3

import csv
import sys
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def convert_csv_to_xlsx(csv_path, xlsx_path=None):
    """Convert CSV to XLSX with proper data types"""
    if xlsx_path is None:
        xlsx_path = csv_path.replace('.csv', '.xlsx')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Telemetry Data"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        headers = next(reader)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_num = 2
        for row in reader:
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_idx)
                header = headers[col_idx - 1].lower()
                
                if header == 'recorded_at':
                    cell.value = value
                    cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                elif header in ['voltage', 'temp']:
                    try:
                        cell.value = float(value)
                        cell.number_format = '0.00'
                    except ValueError:
                        cell.value = value
                elif header == 'is_active':
                    cell.value = value.upper() == 'TRUE'
                    cell.number_format = 'General'
                else:
                    cell.value = value
                    cell.alignment = Alignment(horizontal='left')
            
            row_num += 1
        
        for col_idx, header in enumerate(headers, 1):
            max_length = len(header)
            for row in ws.iter_rows(min_row=2, max_row=row_num-1, min_col=col_idx, max_col=col_idx):
                if row[0].value:
                    max_length = max(max_length, len(str(row[0].value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
    
    wb.save(xlsx_path)
    print(f"Converted {csv_path} to {xlsx_path}")
    return xlsx_path

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: csv_to_xlsx.py <csv_file> [xlsx_file]")
        sys.exit(1)
    
    csv_file = sys.argv[1]
    xlsx_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    if not os.path.exists(csv_file):
        print(f"Error: CSV file not found: {csv_file}")
        sys.exit(1)
    
    convert_csv_to_xlsx(csv_file, xlsx_file)


